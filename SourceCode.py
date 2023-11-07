from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import re
import os
import pandas as pd

def create_excel_sheet(file_path, sheet_name):
    if not os.path.exists(file_path):
        # Create a new workbook and add a sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(['Reg No.','Name','1A01ENG Communicative English','1A02ENG Readings on Kerala','1A07-2MAL Sahithya Ganangal','1A07-2HIN Naya Sahithya','1A07-2ARB Literature in Arabic','1A11BCA Informatics for Computer Applications','1B01BCA Programming in C','1C01MAT-BCA Mathematics for BCA I','Percentage','SGPA','CGPA','Result','Backlogs'])
        # Save the workbook
        workbook.save(file_path)
        print(f"Excel sheet '{file_path}' created successfully.")
    else:
        print(f"Excel sheet '{file_path}' already exists. Skipping creation.")

def ExtractReg(an):
    for row in an.split('\n'):
        if row.startswith(' Reg No.'):
            return row.split()[-1]
        
def ExtractName(an):
    for row in an.split('\n'):
        if row.startswith('Name'):
            return row.split(':')[-1].strip()
        
def ExtractGrade(an,SW):
    for row in an.split('\n'):
        if row.startswith(SW):
            # print(row)
            if row[-6] == 'X':
                return 'X'
            else:
                full=row.split()[-2]
                match = re.search(r'\d+\.\d+',full)
                result = match.group()
                l3c = result[-3:]
                grade = float(l3c)
                if (grade==0.0):
                    if (result[-4]=='1'):
                        return 'A+'
                    else:
                        return 'F'
                elif (grade>=9.0):
                    return 'A+'
                elif (grade>=8.0 and grade<9.0):
                    return 'A'
                elif (grade>=7.0 and grade<8.0):
                    return 'B'
                elif (grade>=6.0 and grade<7.0):
                    return 'C'
                elif (grade>=5.0 and grade<6.0):
                    return 'D'
                elif (grade>=4.0 and grade<5.0):
                    return 'E'
                elif (grade<4.0 and grade>0.0):
                    return 'F'      

def Percentage(an):
    for row in an.split('\n'):
        if row.startswith("Total Marks (%)"):
            pct=row.split()[3]
    if pct=='-':
        return pct
    else:
        return float(pct)
        
def sgpa(an):
    if Percentage(an)=='-':
        return '-'
    else:
        return float(Percentage(an))/10

def cgpa(cg):
    return cg

def PassCheck(an):
    for row in an.split('\n'):
        if row.startswith("Total Marks (%)"):
            l6c=row[-6:]
    return l6c

def BackLogCounter(p1,p2,p31,p32,p33,p4,p5,p6):
    list1 = [p1,p2,p31,p32,p33,p4,p5,p6]
    count=0
    for i in list1:
        if i == 'F' or i =='X':
            count=count+1
    return count

def Enter(folder_path):
    files = os.listdir(folder_path)
    for file in files:
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            # print(file_path.split('\\')[-1])
            print(file_path)
            PDF_URL = file_path #Enter your pdf url here
            # PDF_URL = input("Enter your pdf url ")
            reader = PdfReader(PDF_URL)
            page = reader.pages[0]
            an=page.extract_text()
            # file1 = open("Myfile.txt","w")
            # file1.write(an)
            # file1.close()
            REG_NO = ExtractReg(an)
            NAME = ExtractName(an)
            P1 = ExtractGrade(an,"1A01ENG")
            P2 = ExtractGrade(an,"1A02ENG")
            P3_1 = ExtractGrade(an,"1A07-2HIN")
            P3_2 = ExtractGrade(an,"1A07-2MAL")
            P3_3 = ExtractGrade(an,"1A07-2ARB")
            P4 = ExtractGrade(an,"1A11BCA")
            P5 = ExtractGrade(an,"1B01BCA")
            P6 = ExtractGrade(an,"BCAMathematics")
            PCT = Percentage(an)
            SGPA = sgpa(an)
            CGPA = cgpa(SGPA)
            RESULT = PassCheck(an)
            BACKLOGS = BackLogCounter(P1,P2,P3_1,P3_2,P3_3,P4,P5,P6)
            wb = load_workbook('S1BCA.xlsx')
            ws = wb.active
            empty_row = 1
            while ws.cell(row=empty_row, column=1).value is not None:
                empty_row += 1
            print(empty_row)
            Data = [REG_NO,NAME,P1,P2,P3_1,P3_2,P3_3,P4,P5,P6,PCT,SGPA,CGPA,RESULT,BACKLOGS]
            for col_idx, value in enumerate(Data, start=1):
                cell = ws.cell(row=empty_row, column=col_idx, value=value)
            wb.save('S1BCA.xlsx')

def remove_duplicates(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.drop_duplicates(inplace=True)
    df.to_excel(file_path, sheet_name=sheet_name, index=False)

def sort_sheet(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    first_column_values = [cell.value for cell in sheet['A'][1:]]
    sorted_indexes = sorted(range(2, len(first_column_values) + 2), key=lambda x: first_column_values[x - 2])
    new_sheet = []
    for idx in sorted_indexes:
        new_row = [cell.value for cell in sheet[idx]]
        new_sheet.append(new_row)
    sheet.delete_rows(2, sheet.max_row)
    for r_idx, row in enumerate(new_sheet, start=2):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    workbook.save(file_path)

def center_align(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_col = sheet.max_column
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    workbook.save(file_path)

def col_dimension(file_path, sheet_name):
    # Load the workbook
    workbook = load_workbook(file_path)

    # Select the appropriate sheet
    sheet = workbook[sheet_name]

    # Set the width of the first two columns (A and B)
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20

    # Calculate the total number of cells to be processed
    total_cells = (sheet.max_row - 1) * (sheet.max_column - 2)

    # Save the modified workbook
    workbook.save(file_path)




create_excel_sheet('S1BCA.xlsx', 'Sheet1')
Enter('Reservoir')
remove_duplicates('S1BCA.xlsx', 'Sheet1')
sort_sheet('S1BCA.xlsx', 'Sheet1')
center_align('S1BCA.xlsx', 'Sheet1')
col_dimension('S1BCA.xlsx', 'Sheet1')


input("Excel Sheet Updated Successfully!!!!!\nPress Enter to quit")